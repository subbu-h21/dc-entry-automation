"""
product_matcher.py
==================
Two-stage SKU matching:

Stage 1 — Brand retrieval: narrow ~2000 brand heads to top 20 using
           trigram + phonetic + fuzz.ratio signals, then ask Gemini to
           pick the correct brand.

Stage 2 — SKU pick: score all SKUs under the matched brand with
           fuzz.WRatio, send them to Gemini, return the best match.
"""

import asyncio
import json
import logging
import re

from openai import RateLimitError
from openpyxl import load_workbook
from rapidfuzz import process, fuzz

from config import MATCHING_MODEL
from services.client import get_async_client

log = logging.getLogger(__name__)

_JUNK_PREFIX         = re.compile(r"^([A-Z])\1{3,}")
_JUNK_PREFIX_DISPLAY = re.compile(r"^([A-Za-z])\1{3,}", re.IGNORECASE)
_PUNCT               = re.compile(r"[^A-Z0-9 ]+")


def _normalize(text: str) -> str:
    s = str(text).strip().upper()
    s = _JUNK_PREFIX.sub("", s)
    s = _PUNCT.sub(" ", s)
    return re.sub(r"\s+", " ", s).strip()


def _double_metaphone(word: str) -> tuple[str, str]:
    """
    Pure-Python Double Metaphone phonetic encoding.
    Returns (primary, secondary) codes. No external dependencies.
    """
    word = ''.join(c for c in word.upper() if c.isalpha())
    n = len(word)
    if not n:
        return ('', '')

    pri: list[str] = []
    sec: list[str] = []

    def at(i: int) -> str:
        return word[i] if 0 <= i < n else ''

    def slc(i: int, j: int) -> str:
        return word[max(0, i):max(0, j)]

    def add(p: str, s: str = '') -> None:
        pri.append(p)
        sec.append(s or p)

    pos = 0

    # Skip initial silent pairs
    if slc(0, 2) in ('AE', 'GN', 'KN', 'PN', 'WR'):
        pos = 1

    # Initial vowel encodes as 'A'
    if at(0) in 'AEIOUY':
        add('A')
        pos = 1

    while pos < n:
        c = at(pos)

        if c in 'AEIOUY':
            pos += 1
            continue

        if c == 'B':
            add('P')
            pos += 2 if at(pos + 1) == 'B' else 1

        elif c == 'C':
            if at(pos - 2) not in 'AEIOUY' and slc(pos - 1, pos + 2) == 'ACH' \
                    and at(pos + 2) not in ('I', 'E'):
                add('K'); pos += 2; continue
            if pos == 0 and slc(0, 6) == 'CAESAR':
                add('S'); pos += 2; continue
            if at(pos + 1) == 'H':
                if slc(pos + 2, pos + 4) in ('ET', 'AN', 'EL', 'OU', 'IC', 'IU'):
                    add('K')
                elif slc(0, 4) in ('VAN ', 'VON ') or slc(0, 3) == 'SCH':
                    add('K')
                elif pos == 0 and at(pos + 2) in 'AEIOUY':
                    add('X', 'S')
                else:
                    add('X')
                pos += 2; continue
            if at(pos + 1) == 'Z' and slc(pos - 2, pos + 2) != 'WICZ':
                add('S', 'X'); pos += 2; continue
            if slc(pos + 1, pos + 4) == 'CIA':
                add('X'); pos += 3; continue
            if at(pos + 1) in 'IEY':
                add('S'); pos += 2; continue
            if at(pos + 1) in 'CKQ':
                add('K'); pos += 2; continue
            add('K')
            pos += 1

        elif c == 'D':
            if at(pos + 1) == 'G' and at(pos + 2) in 'IEY':
                add('J'); pos += 3
            elif slc(pos, pos + 2) in ('DT', 'DD'):
                add('T'); pos += 2
            else:
                add('T'); pos += 1

        elif c == 'F':
            add('F')
            pos += 2 if at(pos + 1) == 'F' else 1

        elif c == 'G':
            if at(pos + 1) == 'H':
                if pos > 0 and at(pos - 1) not in 'AEIOUY':
                    add('K'); pos += 2; continue
                if pos == 0:
                    add('J' if at(pos + 2) == 'I' else 'K')
                    pos += 2; continue
                if (pos > 1 and at(pos - 2) in 'BHD') or \
                   (pos > 2 and at(pos - 3) in 'BHD') or \
                   (pos > 3 and at(pos - 4) in 'BH'):
                    pos += 2; continue
                if pos > 2 and at(pos - 1) == 'U' and at(pos - 3) in 'CGLRT':
                    add('F')
                elif pos > 0 and at(pos - 1) != 'I':
                    add('K')
                pos += 2; continue
            if at(pos + 1) == 'N':
                if pos == 1 and at(0) in 'AEIOUY':
                    add('KN', 'N')
                elif slc(pos + 2, pos + 4) == 'EY' or at(pos + 1) == 'Y':
                    add('KN')
                else:
                    add('N', 'KN')
                pos += 2; continue
            if pos == 0 and (at(pos + 1) == 'Y' or slc(pos + 1, pos + 3) in (
                    'ES', 'EP', 'EB', 'EL', 'EY', 'IB', 'IL', 'IN', 'IE', 'EI', 'ER')):
                add('K', 'J'); pos += 2; continue
            if at(pos + 1) in 'IEY':
                add('K', 'J'); pos += 2; continue
            pos += 2 if at(pos + 1) == 'G' else 1
            add('K')

        elif c == 'H':
            if at(pos + 1) in 'AEIOUY' and (pos == 0 or at(pos - 1) in 'AEIOUY'):
                add('H')
            pos += 1

        elif c == 'J':
            if pos == 0 and slc(0, 4) != 'JOSE':
                add('J', 'A')
            elif at(pos - 1) in 'AEIOUY':
                add('J', 'H')
            else:
                add('J')
            pos += 2 if at(pos + 1) == 'J' else 1

        elif c == 'K':
            add('K')
            pos += 2 if at(pos + 1) == 'K' else 1

        elif c == 'L':
            add('L')
            pos += 2 if at(pos + 1) == 'L' else 1

        elif c == 'M':
            add('M')
            pos += 2 if at(pos + 1) == 'M' else 1

        elif c == 'N':
            add('N')
            pos += 2 if at(pos + 1) == 'N' else 1

        elif c == 'P':
            if at(pos + 1) == 'H':
                add('F'); pos += 2
            else:
                add('P')
                pos += 2 if at(pos + 1) == 'P' else 1

        elif c == 'Q':
            add('K')
            pos += 2 if at(pos + 1) == 'Q' else 1

        elif c == 'R':
            add('R')
            pos += 2 if at(pos + 1) == 'R' else 1

        elif c == 'S':
            if slc(pos - 1, pos + 2) in ('ISL', 'YSL'):
                pos += 1; continue
            if pos == 0 and slc(0, 5) == 'SUGAR':
                add('X', 'S'); pos += 1; continue
            if at(pos + 1) == 'H':
                add('X'); pos += 2; continue
            if slc(pos, pos + 3) in ('SIO', 'SIA'):
                add('S', 'X') if at(pos - 1) in 'AEIOUY' else add('X', 'S')
                pos += 3; continue
            if at(pos + 1) == 'C' and at(pos + 2) == 'H':
                if slc(pos + 3, pos + 5) in ('OO', 'ER', 'EN', 'UY', 'ED', 'EM'):
                    add('SK')
                else:
                    add('X')
                pos += 3; continue
            if at(pos + 1) == 'C' and at(pos + 2) in 'IEY':
                add('S'); pos += 3; continue
            add('S')
            pos += 2 if at(pos + 1) in ('S', 'Z') else 1

        elif c == 'T':
            if slc(pos, pos + 3) in ('TIA', 'TIO'):
                add('X'); pos += 3; continue
            if at(pos + 1) == 'H' or slc(pos, pos + 3) == 'TTH':
                if slc(pos + 2, pos + 4) in ('OM', 'AM') or slc(0, 4) in ('VAN ', 'VON '):
                    add('T')
                else:
                    add('T', '0')
                pos += 2; continue
            if at(pos + 1) == 'C' and at(pos + 2) == 'H':
                add('X'); pos += 3; continue
            add('T')
            pos += 2 if at(pos + 1) in ('T', 'D') else 1

        elif c == 'V':
            add('F')
            pos += 2 if at(pos + 1) == 'V' else 1

        elif c == 'W':
            if slc(pos, pos + 2) == 'WR':
                add('R'); pos += 2; continue
            if pos == 0 and (at(pos + 1) in 'AEIOUY' or at(pos + 1) == 'H'):
                add('A', 'F') if at(pos + 1) in 'AEIOUY' else add('A')
                pos += 1; continue
            if slc(pos, pos + 4) in ('WICZ', 'WITZ'):
                add('TS', 'FX'); pos += 4; continue
            pos += 1

        elif c == 'X':
            add('KS')
            pos += 2 if at(pos + 1) in ('C', 'X') else 1

        elif c == 'Z':
            if at(pos + 1) == 'H':
                add('J'); pos += 2; continue
            if slc(pos + 1, pos + 3) in ('ZO', 'ZI', 'ZA'):
                add('S', 'TS')
            else:
                add('S')
            pos += 2 if at(pos + 1) == 'Z' else 1

        else:
            pos += 1

    p = ''.join(pri)
    s = ''.join(sec)
    return (p, s if s != p else p)


def _trigrams(word: str) -> set[str]:
    if len(word) < 3:
        return {word}
    return {word[i:i+3] for i in range(len(word) - 2)}


class ProductMatcher:
    def __init__(self, xlsx_path: str, sheet: str = "data"):
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb[sheet]
        row_iter = ws.iter_rows(values_only=True)
        headers  = [str(c).strip().upper() if c is not None else "" for c in next(row_iter)]
        prod_idx = headers.index("PRODUCT")
        pack_idx = headers.index("PACK")         if "PACK"         in headers else None
        mfr_idx  = headers.index("MANUFACTURER") if "MANUFACTURER" in headers else None

        self.family: dict[str, list] = {}
        self._brand_cache: dict[str, list] = {}

        for i, row in enumerate(row_iter):
            product = str(row[prod_idx]).strip() if row[prod_idx] is not None else ""
            head    = _normalize(product).split()[0] if product else ""
            if not head:
                continue
            self.family.setdefault(head, []).append({
                "id":      i,
                "product": product,
                "pack":    str(row[pack_idx]).strip() if pack_idx is not None and row[pack_idx] is not None else "",
                "mfr":     str(row[mfr_idx]).strip()  if mfr_idx  is not None and row[mfr_idx]  is not None else "",
                "_norm":   _normalize(product),
            })

        wb.close()
        self.heads = list(self.family.keys())

        # Secondary indexes for brand retrieval
        self._trigram_index:  dict[str, set[str]] = {h: _trigrams(h) for h in self.heads}
        self._phonetic_index: dict[str, tuple]    = {h: _double_metaphone(h) for h in self.heads}

    # ------------------------------------------------------------------
    # Stage 1: brand retrieval
    # ------------------------------------------------------------------

    def find_brand(self, first_word: str, top_n: int = 20) -> list[dict]:
        """
        Multi-signal brand head lookup.
        Returns up to top_n candidates ranked by composite score.
        """
        word = _normalize(first_word).split()[0] if first_word else ""
        if not word:
            return []

        if word in self._brand_cache:
            return self._brand_cache[word]

        # Exact match — skip scoring
        if word in self.family:
            result = [{"brand": word, "score": 200, "ratio": 100,
                       "trigram_shared": 0, "phonetic": False, "match_type": "exact"}]
            self._brand_cache[word] = result
            return result

        word_tris     = _trigrams(word)
        word_phonetic = set(_double_metaphone(word)) - {""}

        # Signal 1: trigrams (≥2 shared)
        trigram_hits: dict[str, int] = {}
        for head, head_tris in self._trigram_index.items():
            shared = len(word_tris & head_tris)
            if shared >= 2:
                trigram_hits[head] = shared

        # Signal 2: phonetics (matching Double Metaphone code)
        phonetic_hits: set[str] = set()
        for head, codes in self._phonetic_index.items():
            if word_phonetic & (set(codes) - {""}):
                phonetic_hits.add(head)

        # Signal 3: fuzz.ratio ≥ 45
        ratio_hits: dict[str, float] = {}
        for head in self.heads:
            score = fuzz.ratio(word, head)
            if score >= 45:
                ratio_hits[head] = score

        all_candidates = set(trigram_hits) | phonetic_hits | set(ratio_hits)
        if not all_candidates:
            self._brand_cache[word] = []
            return []

        # Composite score: ratio + (shared_trigrams × 8) + (phonetic_match × 15)
        results = []
        for head in all_candidates:
            ratio    = ratio_hits.get(head, fuzz.ratio(word, head))
            tris     = trigram_hits.get(head, 0)
            phonetic = head in phonetic_hits
            composite = ratio + (tris * 8) + (15 if phonetic else 0)
            results.append({
                "brand":          head,
                "score":          round(composite, 1),
                "ratio":          round(ratio, 1),
                "trigram_shared": tris,
                "phonetic":       phonetic,
                "match_type":     "composite",
            })

        results.sort(key=lambda x: x["score"], reverse=True)
        results = results[:top_n]
        self._brand_cache[word] = results
        return results

    # ------------------------------------------------------------------
    # Stage 2: SKU candidates within a brand family
    # ------------------------------------------------------------------

    def get_sku_candidates(self, brand: str, full_query: str) -> list[dict]:
        """Score all SKUs under a brand by fuzz.WRatio against the full query."""
        if brand not in self.family:
            return []

        family     = self.family[brand]
        query_norm = _normalize(full_query)
        norms      = [r["_norm"] for r in family]
        scored     = process.extract(query_norm, norms, scorer=fuzz.WRatio, limit=None)

        candidates = []
        for pos, (_val, score, idx) in enumerate(scored):
            rec = family[idx]
            candidates.append({
                "index":   pos,
                "id":      rec["id"],
                "product": _JUNK_PREFIX_DISPLAY.sub("", rec["product"]).lstrip(),
                "pack":    rec["pack"],
                "mfr":     rec["mfr"],
                "score":   round(float(score), 1),
            })

        return candidates


# ---------------------------------------------------------------------------
# Gemini: brand pick
# ---------------------------------------------------------------------------

async def ask_gemini_brand(brand_candidates: list[dict], first_word: str) -> dict | None:
    if not brand_candidates:
        return None

    # Exact match — skip LLM call
    if len(brand_candidates) == 1 and brand_candidates[0].get("match_type") == "exact":
        return {**brand_candidates[0], "ai_pick": True, "ai_confidence": 1.0}

    lines = "\n".join(
        f'{i}: {c["brand"]}  (score: {c["score"]})'
        for i, c in enumerate(brand_candidates)
    )
    prompt = (
        f"<request>{first_word}</request>\n\n"
        f"<candidates>\n{lines}\n</candidates>\n\n"
        "<rules>\n"
        "1. Pick the brand name that best matches the requested word.\n"
        "2. Spelling variants and minor OCR errors are acceptable.\n"
        "3. Use pharmaceutical brand knowledge to resolve ambiguity.\n"
        "4. Return -1 if no candidate is a plausible match.\n"
        "</rules>\n\n"
        'Return ONLY valid JSON: {"index": <int>, "confidence": <float>}\n'
        "index = 0-based position in the candidates list, or -1 if no match."
    )

    for attempt in range(3):
        try:
            response = await get_async_client().chat.completions.create(
                model=MATCHING_MODEL,
                temperature=0,
                messages=[{"role": "user", "content": prompt}],
            )
            break
        except RateLimitError:
            if attempt == 2:
                raise
            await asyncio.sleep(2 ** attempt)

    raw = response.choices[0].message.content.strip()

    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        m = re.search(r"\{.*\}", raw, re.S)
        if not m:
            return None
        try:
            data = json.loads(m.group())
        except json.JSONDecodeError:
            return None

    try:
        idx        = int(data["index"])
        confidence = float(data.get("confidence", 0))
    except (KeyError, TypeError, ValueError):
        return None

    if idx == -1:
        return None
    if 0 <= idx < len(brand_candidates):
        return {**brand_candidates[idx], "ai_pick": True, "ai_confidence": confidence}
    return None


# ---------------------------------------------------------------------------
# Gemini: SKU pick
# ---------------------------------------------------------------------------

async def ask_gemini(candidates: list, query: str) -> dict | None:
    if not candidates:
        return None

    if len(candidates) == 1:
        c = candidates[0]
        if _normalize(c["product"]) == query:
            return {"index": 0, "id": c["id"], "product": c["product"], "pack": c["pack"], "confidence": 1.0}

    lines  = "\n".join(
        f'{c["index"]}: {c["product"]}  [{c["pack"]}]  {c["mfr"]}'
        for c in candidates
    )
    prompt = (
        f"<request>{query}</request>\n\n"
        f"<candidates>\n{lines}\n</candidates>\n\n"
        "<rules>\n"
        "1. Dosage form must match. Tablet ≠ Syrup ≠ Drops ≠ Cream ≠ Gel ≠ Injection ≠ Suspension.\n"
        "2. Strength must match. 10mg ≠ 20mg ≠ 40mg ≠ 80mg.\n"
        "3. Combination suffixes mean a DIFFERENT product from the plain molecule.\n"
        "   H=+HCTZ  M=+Metoprolol  AM=+Amlodipine  AT=+Atorvastatin  AV=+Atorvastatin/Vitamin  CT=+Chlorthalidone.\n"
        "   Exception: DS, SR/XL/XR, Forte are NOT combination markers.\n"
        "4. When both plain and combination candidates exist, always pick plain for a plain request.\n"
        "5. Brand name spelling variants are acceptable.\n"
        "6. Do NOT return -1 just because a candidate is the only option.\n"
        "7. Do NOT choose a candidate when it is only a weak spelling coincidence.\n"
        "</rules>\n\n"
        'Return ONLY valid JSON: {"index": <int>, "confidence": <float>}\n'
        "index = 0-based position, or -1 if no reliable match. confidence = 0.0–1.0."
    )

    for attempt in range(3):
        try:
            response = await get_async_client().chat.completions.create(
                model=MATCHING_MODEL,
                temperature=0,
                messages=[{"role": "user", "content": prompt}],
            )
            break
        except RateLimitError:
            if attempt == 2:
                raise
            await asyncio.sleep(2 ** attempt)

    raw = response.choices[0].message.content.strip()

    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        match = re.search(r"\{.*\}", raw, re.S)
        if not match:
            return None
        try:
            data = json.loads(match.group())
        except json.JSONDecodeError:
            return None

    try:
        idx        = int(data["index"])
        confidence = float(data.get("confidence", -1 if idx == -1 else 0))
    except (KeyError, TypeError, ValueError):
        return None

    if idx == -1:
        return {"index": -1, "id": None, "product": None, "pack": None, "confidence": -1, "not_stocked": True}
    if 0 <= idx < len(candidates):
        c = candidates[idx]
        return {"index": idx, "id": c["id"], "product": c["product"], "pack": c["pack"], "confidence": confidence}
    return None


# ---------------------------------------------------------------------------
# Match a whole order
# ---------------------------------------------------------------------------

async def match_order(medicines: list, matcher: ProductMatcher, concurrency: int = 10) -> list:
    sem = asyncio.Semaphore(concurrency)

    async def _resolve(med):
        full_query = _normalize(med["medicine_text"])
        first_word = full_query.split()[0] if full_query else ""

        # Stage 1: brand retrieval
        brand_candidates = matcher.find_brand(first_word)
        async with sem:
            chosen_brand = await ask_gemini_brand(brand_candidates, first_word)

        if not chosen_brand:
            return {
                **med,
                "match_type":       "none",
                "family":           None,
                "brand_confidence": None,
                "candidates_count": 0,
                "top_candidates":   [],
                "gemini_pick":      None,
                "matched_id":       None,
                "matched_product":  None,
                "matched_pack":     None,
                "match_confidence": None,
                "not_stocked":      False,
            }

        brand = chosen_brand["brand"]

        # Stage 2: SKU matching within brand family
        sku_candidates = matcher.get_sku_candidates(brand, med["medicine_text"])
        async with sem:
            chosen_sku = await ask_gemini(sku_candidates, full_query)

        # Top 10 for user correction panel
        top_candidates = [dict(c) for c in sku_candidates[:10]]
        gemini_pick_extra = None

        if chosen_sku and not chosen_sku.get("not_stocked"):
            gemini_idx = chosen_sku.get("index")
            in_top = False
            for c in top_candidates:
                if c["index"] == gemini_idx:
                    c["ai_pick"] = True
                    in_top = True
                else:
                    c["ai_pick"] = False
            if not in_top:
                gemini_pick_extra = {**chosen_sku, "ai_pick": True}
        else:
            for c in top_candidates:
                c["ai_pick"] = False

        return {
            **med,
            "match_type":       chosen_brand.get("match_type", "composite"),
            "family":           brand,
            "brand_confidence": chosen_brand.get("ai_confidence"),
            "candidates_count": len(sku_candidates),
            "top_candidates":   top_candidates,
            "gemini_pick":      gemini_pick_extra,
            "matched_id":       chosen_sku["id"]      if chosen_sku and not chosen_sku.get("not_stocked") else None,
            "matched_product":  chosen_sku["product"] if chosen_sku and not chosen_sku.get("not_stocked") else None,
            "matched_pack":     chosen_sku["pack"]    if chosen_sku and not chosen_sku.get("not_stocked") else None,
            "match_confidence": chosen_sku.get("confidence") if chosen_sku else None,
            "not_stocked":      bool(chosen_sku and chosen_sku.get("not_stocked")),
        }

    results = list(await asyncio.gather(*[_resolve(med) for med in medicines]))
    log.info("Matched %d/%d medicines", sum(1 for r in results if r.get("matched_product")), len(results))
    return results
